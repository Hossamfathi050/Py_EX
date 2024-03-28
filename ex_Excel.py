import openpyxl as xl
from openpyxl.chart import bar_chart,Reference

wb=xl.load_workbook('excel1.xlsx')
sheet=wb['h']
cell=sheet.cell(2,2)
print(cell.value)
print(sheet.max_row)
for row in range(2,sheet.max_row+1):
    cell=sheet.cell(row,3)
    newprice=cell.value*3
    newprice_cell=sheet.cell(row,4)
    newprice_cell.value=newprice

wb.save('excel1.xlsx')    


values=Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)
chart=bar_chart()
chart.add_data
